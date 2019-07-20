# 加载数据分析常用库
from openpyxl import load_workbook
from openpyxl import Workbook
import numpy as np
import tensorflow as tf
# from sklearn.metrics import mean_absolute_error, mean_squared_error
import matplotlib.pyplot as plt
import scipy.io as scio
import warnings

book = load_workbook(filename=r"E:\PHD\0论文\0论文\TVT_2018\code\time series prediction_CSDN\multi-point prediction online\prediction_360_length_5.xlsx")
sheetnames = book.get_sheet_names()
sheet = book.get_sheet_by_name(sheetnames[0])

wb = Workbook()
ws = wb.active

dataFile = 'channel_data_360_10.mat'
data_com = scio.loadmat(dataFile)
data_real = np.abs(data_com['h'])
y = data_real[0][:8000]

warnings.filterwarnings('ignore')

# 定义常量（答辩状主要展示下面两个参数的影响）
time_step = 5
iter_time = 200
iter_time_new = 1  # 滑动训练，一次训练一次

# 定义常量
rnn_unit = 5  # hidden layer units
input_size = 1
output_size = 1
train_end = 4800
data_num = len(y)
lr = 0.03  # 学习率
batch_size = None  # 因为数据量较小，所以全部用，不分批
train_begin = 0
validation_end = 6400

tf.reset_default_graph()

# 输入层、输出层权重、偏置
weights = {
    'in': tf.Variable(tf.random_normal([input_size, rnn_unit])),
    'out': tf.Variable(tf.random_normal([rnn_unit, 1]))
}
biases = {
    'in': tf.Variable(tf.constant(0.1, shape=[rnn_unit, ])),
    'out': tf.Variable(tf.constant(0.1, shape=[1, ]))
}


def get_data(time_step, train_begin, train_end, validation_end):
    data_m, data_validation_m, data_test_m, data_total_m = [], [], [], []

    # 这个地方需要减一，“1”意味着为最后的预测腾出一个位置
    for i in range(train_end - time_step):
        data_m.append(y[i:i + time_step])

    # 这里的维度需要再斟酌一下
    data_x = np.reshape(data_m, [-1, time_step])
    data_train_x = data_x[:, :, np.newaxis]
    data_train_y = y[train_begin + time_step:train_end]
    data_train_y = np.reshape(data_train_y, [-1, 1])
    # data_train_y = data_y[:, :, np.newaxis]

    # 分批处理
    # for i in range(np.shape(data_x)[0] - batch_size):
    #     if i % batch_size == 0:
    #         batch_index.append(i)
    #     data_m_x.append(data_x[i:i + batch_size, :input_size])
    #     data_m_y.append(data_y[i:i + batch_size, np.newaxis])
    #
    # data_train_x = np.reshape(data_m_x, [-1, time_step, input_size])
    # data_train_y = np.reshape(data_m_y, [-1, time_step, output_size])

    for i in range(train_end - time_step, validation_end - time_step):
        data_validation_m.append(y[i:i + time_step])
    data_validation_x = np.reshape(data_validation_m, [-1, time_step])
    data_validation_x = data_validation_x[:, :, np.newaxis]
    data_validation_y = y[train_end:validation_end]
    data_validation_y = np.reshape(data_validation_y, [-1, 1])

    for i in range(validation_end - time_step, len(y) - time_step):
        data_test_m.append(y[i:i + time_step])
    data_test_x = np.reshape(data_test_m, [-1, time_step])
    data_test_x = data_test_x[:, :, np.newaxis]
    data_test_y = y[validation_end:len(y)]
    data_test_y = np.reshape(data_test_y, [-1, 1])

    # 构造总数据，为滑动训练更新作准备
    for i in range(len(y) - time_step):
        data_total_m.append(y[i:i + time_step])
    data_total_x = np.reshape(data_total_m, [-1, time_step])
    data_total_x = data_total_x[:, :, np.newaxis]
    data_total_y = y[time_step:len(y)]
    data_total_y = np.reshape(data_total_y, [-1, 1])

    # data_test_m_x_m, data_test_m_y_m = [], []
    # for i in range(np.shape(data_test_m_x)[0] - time_step):
    #     data_test_m_x_m.append(data_test_m_x[i:i+time_step, :input_size])
    #     data_test_m_y_m.append(data_test_m_y[i:i+time_step, np.newaxis])
    #
    # data_test_x = np.reshape(data_test_m_x_m, [-1, time_step, input_size])
    # data_test_y = np.reshape(data_test_m_y_m, [-1, time_step, output_size])

    return data_train_x, data_train_y, data_validation_x, data_validation_y, data_test_x, data_test_y, data_total_x, data_total_y


def lstm(X):
    batch_size = tf.shape(X)[0]
    time_step = tf.shape(X)[1]
    w_in = weights['in']
    b_in = biases['in']
    input = tf.reshape(X, [-1, input_size])  # 需要将tensor转成2维进行计算，计算后的结果作为隐藏层的输入
    input_rnn = tf.matmul(input, w_in) + b_in
    input_rnn = tf.reshape(input_rnn, [-1, time_step, rnn_unit])  # 将tensor转成3维，作为lstm cell的输入
    cell = tf.contrib.rnn.BasicLSTMCell(rnn_unit)
    # cell=tf.contrib.rnn.core_rnn_cell.BasicLSTMCell(rnn_unit)
    init_state = cell.zero_state(batch_size, dtype=tf.float32)
    output_rnn, final_states = tf.nn.dynamic_rnn(cell, input_rnn, initial_state=init_state, dtype=tf.float32)
    print(output_rnn)
    print(final_states)
    # output_rnn是记录lstm每个输出节点的结果，final_states是最后一个cell的结果
    output = tf.reshape(output_rnn, [-1, rnn_unit])
    final_output = final_states.h
    final_output = tf.reshape(final_output, [-1, rnn_unit])
    # 作为输出层的输入
    w_out = weights['out']
    b_out = biases['out']
    pred = tf.matmul(final_output, w_out) + b_out
    # pred = tf.add(pred, 0, name="pred1")
    # tf.add_to_collection(name='pred', value=pred)
    return pred, final_states


def train_lstm(X, Y):
    pred, _ = lstm(X)
    # 损失函数
    # 这是将数据变成了一列
    loss = tf.reduce_mean(tf.square(tf.reshape(pred, [-1]) - tf.reshape(Y, [-1])))
    # with tf.name_scope("train1"):
    train_op = tf.train.AdamOptimizer(lr).minimize(loss)
    # tf.add_to_collection(name='train1', value=train_op)
    # train_op = tf.add(train_op, 0, name="train1")
    return pred, loss, train_op


X = tf.placeholder(tf.float32, shape=[None, time_step, input_size])
Y = tf.placeholder(tf.float32, shape=[None, output_size])

train_x, train_y, validation_x, validation_y, test_x, test_y, total_x, total_y = get_data(time_step, train_begin, train_end, validation_end)

with tf.Session() as sess:
    pred, loss, train_op = train_lstm(X, Y)
    sess.run(tf.global_variables_initializer())
    # 先重复训练
    train_loss_return = []
    validation_loss_return = np.zeros(iter_time)
    test_loss_return = []

    for i in range(iter_time):
        _, train_loss = sess.run([train_op, loss], feed_dict={X: train_x, Y: train_y})
        validation_loss = sess.run(loss, feed_dict={X: validation_x, Y: validation_y})
        test_loss = sess.run(loss, feed_dict={X: test_x, Y: test_y})
        train_loss_return.append(train_loss)
        validation_loss_return[i] = validation_loss
        test_loss_return.append(0)
        print('iter:', i, 'train_loss:', train_loss, 'validation_loss', validation_loss, 'test_loss', test_loss)

    # 求出在预训练模型上的表现情况
    train_predict_pre_training = sess.run(pred, feed_dict={X: train_x})
    train_predict_pre_training = train_predict_pre_training.reshape((-1))

    validation_predict_pre_training = sess.run(pred, feed_dict={X: validation_x})
    validation_predict_pre_training = validation_predict_pre_training.reshape((-1))


    # 保存模型
#     saver = tf.train.Saver()
#     saver.save(sess, "save_net/net.ckpt")
#
# with tf.Session() as sess:
#     saver = tf.train.import_meta_graph("save_net/net.ckpt.meta")
#     saver.restore(sess, tf.train.latest_checkpoint("save_net"))
#     graph = tf.get_default_graph()
#     X = graph.get_tensor_by_name("X:0")
#     Y = graph.get_tensor_by_name("Y:0")
#     pred = graph.get_tensor_by_name("pred1:0")
#     train_op = tf.get_collection('train1')
#     # graph.get_operation_by_name()

    test_predict = []
    train_new_size = 100

    prediction_length1 = 5

    # 先预测下一个时刻的值
    test_x_buff = total_x[validation_end - time_step]
    # test_x_buff1 = np.reshape(test_x_buff, [-1])
    # test_x_buff1 = test_x_buff1[np.newaxis, :, np.newaxis]
    test_x_buff = test_x_buff[np.newaxis, :, :]
    # print(test_x_buff == test_x_buff1)
    for i in range(1, prediction_length1):
        test_predict_result_1_buff = sess.run([pred], feed_dict={X: test_x_buff})
        # test_loss_return.append(test_loss)
        test_predict.append(test_predict_result_1_buff)
        test_x_buff = total_x[validation_end - time_step + i]
        test_x_buff = np.reshape(test_x_buff, [-1])
        test_x_buff[time_step-1] = np.reshape(test_predict_result_1_buff, [-1])[0]
        test_x_buff = test_x_buff[np.newaxis, :, np.newaxis]
    test_predict_result_1_buff = sess.run(pred, feed_dict={X: test_x_buff})
    test_predict.append(test_predict_result_1_buff)
    # test_predict_result_1_buff = sess.run(pred, feed_dict={X: test_x_buff})
    # test_predict.append(test_predict_result_1_buff)
    # test_x_buff = total_x[validation_end - time_step + 2]
    # test_x_buff = np.reshape(test_x_buff, [-1])
    # test_x_buff[4] = np.reshape(test_predict_result_1_buff, [-1])[0]
    # test_x_buff = test_x_buff[np.newaxis, :, np.newaxis]
    # test_predict_result_1_buff = sess.run(pred, feed_dict={X: test_x_buff})
    # test_predict.append(test_predict_result_1_buff)
    # test_x_buff = total_x[validation_end - time_step + 3]
    # test_x_buff = np.reshape(test_x_buff, [-1])
    # test_x_buff[4] = np.reshape(test_predict_result_1_buff, [-1])[0]
    # test_x_buff = test_x_buff[np.newaxis, :, np.newaxis]
    # test_predict_result_1_buff = sess.run(pred, feed_dict={X: test_x_buff})
    # test_predict.append(test_predict_result_1_buff)
    # test_x_buff = total_x[validation_end - time_step + 4]
    # test_x_buff = np.reshape(test_x_buff, [-1])
    # test_x_buff[4] = np.reshape(test_predict_result_1_buff, [-1])[0]
    # test_x_buff = test_x_buff[np.newaxis, :, np.newaxis]
    # test_predict_result_1_buff = sess.run(pred, feed_dict={X: test_x_buff})
    # test_predict.append(test_predict_result_1_buff)

    for i in range(validation_end - time_step + prediction_length1, data_num - time_step, prediction_length1):

        # 对于新数据进行新的训练（部分样本学习【只学习最近更新的】),往后滑动一个
        train_x_new = total_x[i - train_new_size:i]
        train_y_new = total_y[i - train_new_size:i]
        for j in range(iter_time_new):
            _, train_loss = sess.run([train_op, loss], feed_dict={X: train_x_new, Y: train_y_new})
            train_loss_return.append(train_loss)

        # 预测下一个时刻的值
        test_x_buff = total_x[i]
        test_x_buff = test_x_buff[np.newaxis, :, :]
        for j in range(1, prediction_length1):
            test_predict_result_1_buff = sess.run([pred], feed_dict={X: test_x_buff})
            # test_loss_return.append(test_loss)
            test_predict.append(test_predict_result_1_buff)
            test_x_buff = total_x[i + j]
            test_x_buff = np.reshape(test_x_buff, [-1])
            test_x_buff[time_step-1] = np.reshape(test_predict_result_1_buff, [-1])[0]
            test_x_buff = test_x_buff[np.newaxis, :, np.newaxis]
        test_predict_result_1_buff = sess.run(pred, feed_dict={X: test_x_buff})
        test_predict.append(test_predict_result_1_buff)
        # test_predict_result_1_buff = sess.run(pred, feed_dict={X: test_x_buff})
        # test_predict.append(test_predict_result_1_buff)
        # test_x_buff = total_x[i + 2]
        # test_x_buff = np.reshape(test_x_buff, [-1])
        # test_x_buff[4] = np.reshape(test_predict_result_1_buff, [-1])[0]
        # test_x_buff = test_x_buff[np.newaxis, :, np.newaxis]
        # test_predict_result_1_buff = sess.run(pred, feed_dict={X: test_x_buff})
        # test_predict.append(test_predict_result_1_buff)
        # test_x_buff = total_x[i + 3]
        # test_x_buff = np.reshape(test_x_buff, [-1])
        # test_x_buff[4] = np.reshape(test_predict_result_1_buff, [-1])[0]
        # test_x_buff = test_x_buff[np.newaxis, :, np.newaxis]
        # test_predict_result_1_buff = sess.run(pred, feed_dict={X: test_x_buff})
        # test_predict.append(test_predict_result_1_buff)
        # test_x_buff = total_x[i + 4]
        # test_x_buff = np.reshape(test_x_buff, [-1])
        # test_x_buff[4] = np.reshape(test_predict_result_1_buff, [-1])[0]
        # test_x_buff = test_x_buff[np.newaxis, :, np.newaxis]
        # test_predict_result_1_buff = sess.run(pred, feed_dict={X: test_x_buff})
        # test_predict.append(test_predict_result_1_buff)

    test_predict = np.reshape(test_predict, [-1])
    # test_predict = test_predict.reshape((-1))

    train_predict = sess.run(pred, feed_dict={X: train_x})
    train_predict = train_predict.reshape((-1))

    total_predict = sess.run(pred, feed_dict={X: total_x})
    total_predict = total_predict.reshape((-1))

    validation_predict = sess.run(pred, feed_dict={X: validation_x})
    validation_predict = validation_predict.reshape((-1))

for i in range(len(train_loss_return)):
        ws.cell(row=i + 1, column=1).value = train_loss_return[i]
for i in range(len(validation_loss_return)):
        ws.cell(row=i + 1, column=2).value = validation_loss_return[i]
for i in range(len(test_loss_return)):
        ws.cell(row=i + 1, column=3).value = test_loss_return[i]

for i in range(len(train_predict)):
        ws.cell(row=i + 1, column=4).value = train_predict[i]

for i in range(len(validation_predict)):
        ws.cell(row=i + 1, column=5).value = validation_predict[i]

for i in range(len(total_predict)):
        ws.cell(row=i + 1, column=6).value = total_predict[i]
# 增量学习时的测试集表现
for i in range(len(test_predict)):
        ws.cell(row=i + 1, column=7).value = test_predict[i]
# 预训练集上的表现
for i in range(len(train_predict_pre_training)):
        ws.cell(row=i + 1, column=8).value = train_predict_pre_training[i]
# 预验证集上的表现
for i in range(len(validation_predict_pre_training)):
        ws.cell(row=i + 1, column=9).value = validation_predict_pre_training[i]


wb.save(filename="prediction_360_length_5.xlsx")

plt.figure(figsize=(24, 8))
plt.plot(y[:-1])
plt.plot([None for _ in range(train_end)] + [None for _ in range(train_end, validation_end)] + [x for x in test_predict])
plt.plot([m for m in train_predict] + [None for _ in range(train_end, data_num)])
plt.plot([None for _ in range(train_end)] + [n for n in validation_predict] + [None for _ in range(validation_end, data_num)])
plt.plot([m for m in train_predict_pre_training] + [None for _ in range(train_end, data_num)])
plt.plot([None for _ in range(train_end)] + [n for n in validation_predict_pre_training] + [None for _ in range(validation_end, data_num)])
# plt.plot([k for k in total_predict])
plt.legend(labels=['Real_data', 'Prediction on test', 'Prediction on train', 'Prediction on validation', 'Prediction on train_pre', 'Prediction on validation_pre'])
plt.show()

plt.figure()
plt.plot(train_loss_return[:-1])
plt.plot(validation_loss_return[:-1])
plt.plot(test_loss_return[:-1])
plt.legend(labels=['Loss on train', 'Loss on validation', 'Loss on test'])
plt.show()

